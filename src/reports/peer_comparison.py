import re
import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


OUTPUT_PATH = Path("output/peer_comparison.xlsx")

METRIC_COLUMNS = [
    "return_on_equity_pct",
    "return_on_capital_employed_pct",
    "net_profit_margin_pct",
    "operating_profit_margin_pct",
    "debt_to_equity",
    "interest_coverage",
    "asset_turnover",
    "free_cash_flow_cr",
    "capex_cr",
    "earnings_per_share",
    "book_value_per_share",
    "dividend_payout_ratio_pct",
    "total_debt_cr",
    "cash_from_operations_cr",
    "revenue_cagr_5yr",
    "pat_cagr_5yr",
    "eps_cagr_5yr",
    "composite_quality_score",
    "pe_ratio",
    "pb_ratio",
]

GREEN_FILL = "C6EFCE"
YELLOW_FILL = "FFEB9C"
RED_FILL = "FFC7CE"
BENCHMARK_FILL = "F4B183"
SUMMARY_FILL = "D9EAF7"


def generate_peer_comparison_report(
    db_path="db/nifty100.db",
    output_path=OUTPUT_PATH,
):
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    data = load_peer_comparison_data(db_path)
    worksheets = build_peer_group_frames(data)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for peer_group_name, frame in worksheets.items():
            sheet_name = _sheet_name(peer_group_name)
            frame.to_excel(writer, sheet_name=sheet_name, index=False)

    _format_workbook(output_path, worksheets)
    return output_path


def load_peer_comparison_data(db_path):
    with sqlite3.connect(db_path) as connection:
        financial_ratios = pd.read_sql_query("SELECT * FROM financial_ratios", connection)
        companies = pd.read_sql_query(
            """
            SELECT company_id, company_name, roce_percentage
            FROM companies
            """,
            connection,
        )
        peer_groups = pd.read_sql_query(
            """
            SELECT peer_group_name, company_id, is_benchmark
            FROM peer_groups
            """,
            connection,
        )
        peer_percentiles = pd.read_sql_query(
            "SELECT * FROM peer_percentiles",
            connection,
        )
        market_cap = (
            pd.read_sql_query("SELECT * FROM market_cap", connection)
            if _table_exists(connection, "market_cap")
            else pd.DataFrame()
        )

    if "return_on_capital_employed_pct" not in financial_ratios.columns:
        financial_ratios = financial_ratios.merge(
            companies[["company_id", "roce_percentage"]],
            on="company_id",
            how="left",
        ).rename(columns={"roce_percentage": "return_on_capital_employed_pct"})

    return {
        "financial_ratios": financial_ratios,
        "companies": companies[["company_id", "company_name"]],
        "peer_groups": peer_groups,
        "peer_percentiles": peer_percentiles,
        "market_cap": market_cap,
    }


def build_peer_group_frames(data):
    financial_ratios = _latest_rows(data["financial_ratios"])
    market_cap = _latest_rows(data["market_cap"]) if not data["market_cap"].empty else pd.DataFrame()
    peer_groups = data["peer_groups"].copy()

    if not market_cap.empty:
        financial_ratios = financial_ratios.merge(
            market_cap[["company_id", "pe_ratio", "pb_ratio"]],
            on="company_id",
            how="left",
        )

    report_source = (
        peer_groups.merge(data["companies"], on="company_id", how="left")
        .merge(financial_ratios, on="company_id", how="left")
        .copy()
    )

    peer_percentiles = data["peer_percentiles"].copy()
    frames = {}
    for peer_group_name, group in report_source.groupby("peer_group_name", sort=True):
        group = group.sort_values(["is_benchmark", "company_id"], ascending=[False, True])
        frames[peer_group_name] = _build_peer_group_frame(
            peer_group_name,
            group,
            peer_percentiles,
        )
    return frames


def _build_peer_group_frame(peer_group_name, group, peer_percentiles):
    rows = []
    for _, row in group.iterrows():
        output_row = {
            "company_id": row["company_id"],
            "company_name": row.get("company_name"),
        }

        year = row.get("year")
        for metric in METRIC_COLUMNS:
            output_row[metric] = _numeric_value(row.get(metric))
            output_row[f"{metric}_percentile_rank"] = _percentile_value(
                peer_percentiles,
                row["company_id"],
                peer_group_name,
                metric,
                year,
                group,
            )

        rows.append(output_row)

    frame = pd.DataFrame(rows)
    median_row = {"company_id": "MEDIAN", "company_name": "Peer Group Median"}
    for metric in METRIC_COLUMNS:
        median_row[metric] = pd.to_numeric(frame[metric], errors="coerce").median()
        median_row[f"{metric}_percentile_rank"] = None

    frame = pd.concat([frame, pd.DataFrame([median_row])], ignore_index=True)
    return frame


def _percentile_value(peer_percentiles, company_id, peer_group_name, metric, year, group):
    if pd.notna(year):
        matches = peer_percentiles[
            peer_percentiles["company_id"].eq(company_id)
            & peer_percentiles["peer_group_name"].eq(peer_group_name)
            & peer_percentiles["metric"].eq(metric)
            & peer_percentiles["year"].eq(int(year))
        ]
        if not matches.empty:
            return round(float(matches["percentile_rank"].iloc[0]) * 100, 2)

    values = pd.to_numeric(group[metric], errors="coerce") if metric in group.columns else pd.Series(dtype=float)
    value = _numeric_value(group[group["company_id"].eq(company_id)][metric].iloc[0]) if metric in group.columns else None
    percentile = _percent_rank_for_value(values, value, invert=metric == "debt_to_equity")
    return round(percentile * 100, 2) if percentile is not None else None


def _percent_rank_for_value(values, value, invert=False):
    if value is None or pd.isna(value):
        return None

    values = values.dropna()
    if values.empty:
        return None
    if len(values) == 1:
        percentile = 0.0
    else:
        rank = values.rank(method="min", ascending=True).loc[values[values.eq(value)].index[0]]
        percentile = float((rank - 1) / (len(values) - 1))
    if invert:
        percentile = 1 - percentile
    return max(0.0, min(percentile, 1.0))


def _latest_rows(frame):
    if frame.empty or not {"company_id", "year"}.issubset(frame.columns):
        return frame.copy()

    latest = frame.dropna(subset=["company_id", "year"]).copy()
    latest["year"] = pd.to_numeric(latest["year"], errors="coerce")
    return (
        latest.sort_values(["company_id", "year"])
        .drop_duplicates(subset=["company_id"], keep="last")
        .reset_index(drop=True)
    )


def _format_workbook(output_path, worksheets):
    workbook = load_workbook(output_path)
    green = PatternFill("solid", fgColor=GREEN_FILL)
    yellow = PatternFill("solid", fgColor=YELLOW_FILL)
    red = PatternFill("solid", fgColor=RED_FILL)
    benchmark = PatternFill("solid", fgColor=BENCHMARK_FILL)
    summary = PatternFill("solid", fgColor=SUMMARY_FILL)

    for peer_group_name, frame in worksheets.items():
        worksheet = workbook[_sheet_name(peer_group_name)]
        headers = {cell.value: cell.column for cell in worksheet[1]}
        percentile_columns = [
            column
            for column, index in headers.items()
            if isinstance(column, str) and column.endswith("_percentile_rank")
        ]

        last_data_row = worksheet.max_row - 1
        for column_name in percentile_columns:
            column_letter = get_column_letter(headers[column_name])
            cell_range = f"{column_letter}2:{column_letter}{last_data_row}"
            worksheet.conditional_formatting.add(
                cell_range,
                CellIsRule(operator="greaterThanOrEqual", formula=["75"], fill=green),
            )
            worksheet.conditional_formatting.add(
                cell_range,
                CellIsRule(operator="between", formula=["25.000001", "74.999999"], fill=yellow),
            )
            worksheet.conditional_formatting.add(
                cell_range,
                CellIsRule(operator="lessThanOrEqual", formula=["25"], fill=red),
            )

        benchmark_ids = set(
            data_company_id
            for data_company_id in frame.loc[
                frame["company_id"].ne("MEDIAN"), "company_id"
            ].head(1)
        )
        for row_index in range(2, worksheet.max_row):
            company_id = worksheet.cell(row=row_index, column=headers["company_id"]).value
            if company_id in benchmark_ids:
                for cell in worksheet[row_index]:
                    cell.fill = benchmark
                    cell.font = Font(bold=True)

        for cell in worksheet[worksheet.max_row]:
            cell.fill = summary
            cell.font = Font(bold=True)

        worksheet.freeze_panes = "C2"
        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 28)

    workbook.save(output_path)


def _numeric_value(value):
    if pd.isna(value):
        return None
    return float(value)


def _sheet_name(name):
    cleaned = re.sub(r"[\[\]:*?/\\]", " ", str(name)).strip()
    return cleaned[:31] or "Peer Group"


def _table_exists(connection, table_name):
    return (
        connection.execute(
            """
            SELECT COUNT(*)
            FROM sqlite_master
            WHERE type = 'table'
              AND name = ?
            """,
            (table_name,),
        ).fetchone()[0]
        == 1
    )


if __name__ == "__main__":
    path = generate_peer_comparison_report()
    print(f"Generated peer comparison report: {path}")
