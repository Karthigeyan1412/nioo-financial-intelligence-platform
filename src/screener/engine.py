from pathlib import Path
import sqlite3

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pandas as pd
import yaml

from src.analytics.cagr import calculate_metric_cagrs


DEFAULT_CONFIG_PATH = "screener_config.yaml"
FINANCIAL_SECTOR = "FINANCIALS"
GREEN_FILL = "C6EFCE"
RED_FILL = "FFC7CE"


SCORING_METRICS = [
    ("return_on_equity_pct", 0.15, True),
    ("return_on_capital_employed_pct", 0.10, True),
    ("net_profit_margin_pct", 0.10, True),
    ("fcf_cagr_5yr", 0.15, True),
    ("cfo_pat_ratio", 0.10, True),
    ("positive_free_cash_flow_flag", 0.05, True),
    ("revenue_cagr_5yr", 0.10, True),
    ("pat_cagr_5yr", 0.10, True),
    ("debt_to_equity", 0.10, False),
    ("interest_coverage", 0.05, True),
]

EXPORT_COLUMNS = [
    ("company_name", "Company"),
    ("company_id", "Company ID"),
    ("broad_sector", "Sector"),
    ("year", "Year"),
    ("return_on_equity_pct", "ROE"),
    ("return_on_capital_employed_pct", "ROCE"),
    ("net_profit_margin_pct", "Net Profit Margin"),
    ("operating_profit_margin_pct", "Operating Profit Margin"),
    ("revenue_cagr_5yr", "Revenue CAGR"),
    ("pat_cagr_5yr", "PAT CAGR"),
    ("eps_cagr_5yr", "EPS CAGR"),
    ("debt_to_equity", "Debt-to-Equity"),
    ("interest_coverage", "Interest Coverage"),
    ("free_cash_flow_cr", "Free Cash Flow"),
    ("fcf_cagr_5yr", "FCF CAGR"),
    ("cfo_pat_ratio", "CFO/PAT Ratio"),
    ("pe_ratio", "P/E"),
    ("pb_ratio", "P/B"),
    ("dividend_yield_pct", "Dividend Yield"),
    ("dividend_payout_ratio_pct", "Dividend Payout"),
    ("market_cap_crore", "Market Cap"),
    ("sales", "Sales"),
    ("net_profit", "Net Profit"),
    ("asset_turnover", "Asset Turnover"),
    ("composite_quality_score", "Composite Quality Score"),
    ("sector_relative_score", "Sector Relative Score"),
]


PRESETS = {
    "quality_compounder": {
        "roe_min": 15,
        "debt_to_equity_max": 1,
        "free_cash_flow_min": 0,
        "revenue_cagr_5yr_min": 10,
    },
    "value_pick": {
        "pe_max": 20,
        "pb_max": 3,
        "debt_to_equity_max": 2,
        "dividend_yield_min": 1,
    },
    "growth_accelerator": {
        "pat_cagr_5yr_min": 20,
        "revenue_cagr_5yr_min": 15,
        "debt_to_equity_max": 2,
    },
    "dividend_champion": {
        "dividend_yield_min": 2,
        "dividend_payout_max": 80,
        "free_cash_flow_min": 0,
    },
    "debt_free_blue_chip": {
        "debt_to_equity_eq": 0,
        "roe_min": 12,
        "sales_min": 5000,
    },
    "turnaround_watch": {
        "revenue_cagr_3yr_min": 10,
        "free_cash_flow_min": 0,
        "debt_to_equity_declining": True,
    },
}


class ScreenerEngine:
    def __init__(self, config_path=DEFAULT_CONFIG_PATH):
        self.config_path = Path(config_path)
        self.config = load_screener_config(self.config_path)

    def apply_filters(self, financial_ratios, filters):
        result = ensure_composite_quality_score(financial_ratios.copy())

        for filter_name, threshold in (filters or {}).items():
            result = self._apply_filter(result, filter_name, threshold)

        return result.sort_values(
            "composite_quality_score",
            ascending=False,
            na_position="last",
        ).reset_index(drop=True)

    def run_preset(self, preset_name, financial_ratios):
        return self.apply_filters(financial_ratios, PRESETS[preset_name])

    def _apply_filter(self, df, filter_name, threshold):
        if filter_name == "debt_to_equity_eq":
            return _apply_debt_to_equity_equal(df, threshold)

        if filter_name == "debt_to_equity_declining":
            if not threshold:
                return df
            if "debt_to_equity_declining" not in df.columns:
                df = add_debt_to_equity_declining(df)
            return df[df["debt_to_equity_declining"].fillna(False)]

        metric = self.config["metrics"].get(filter_name)
        if metric is None:
            raise KeyError(f"Unknown screener filter: {filter_name}")

        column = metric["column"]
        operator = metric["operator"]

        if column not in df.columns:
            return df.iloc[0:0]

        if filter_name == "debt_to_equity_max":
            return _apply_debt_to_equity_max(df, column, threshold)

        if filter_name == "interest_coverage_min":
            values = _effective_interest_coverage(df)
        else:
            values = pd.to_numeric(df[column], errors="coerce")

        if operator == "min":
            return df[values > threshold]

        if operator == "max":
            return df[values < threshold]

        raise ValueError(f"Unsupported operator for {filter_name}: {operator}")


def load_screener_config(config_path=DEFAULT_CONFIG_PATH):
    with Path(config_path).open("r", encoding="utf-8") as file:
        return yaml.safe_load(file)


def load_screener_dataset(db_path="db/nifty100.db"):
    with sqlite3.connect(db_path) as connection:
        ratios = pd.read_sql_query("SELECT * FROM financial_ratios", connection)
        market_cap = pd.read_sql_query("SELECT * FROM market_cap", connection)
        profitandloss = pd.read_sql_query("SELECT * FROM profitandloss", connection)
        cashflow = pd.read_sql_query("SELECT * FROM cashflow", connection)
        sectors = pd.read_sql_query("SELECT company_id, broad_sector FROM sectors", connection)
        companies = pd.read_sql_query(
            "SELECT company_id, company_name, roce_percentage FROM companies",
            connection,
        )

    dataset = ratios.merge(
        market_cap[
            [
                "company_id",
                "year",
                "market_cap_crore",
                "pe_ratio",
                "pb_ratio",
                "dividend_yield_pct",
            ]
        ],
        on=["company_id", "year"],
        how="left",
    )
    dataset = dataset.merge(
        profitandloss[["company_id", "year", "sales", "net_profit"]],
        on=["company_id", "year"],
        how="left",
    )
    dataset = dataset.merge(
        cashflow[["company_id", "year", "operating_activity", "investing_activity"]],
        on=["company_id", "year"],
        how="left",
    )
    dataset = dataset.merge(sectors, on="company_id", how="left")
    dataset = dataset.merge(companies, on="company_id", how="left")
    dataset["return_on_capital_employed_pct"] = pd.to_numeric(
        dataset.get("return_on_capital_employed_pct", dataset["roce_percentage"]),
        errors="coerce",
    )
    dataset["cfo_pat_ratio"] = _cfo_pat_ratio(dataset)
    dataset["positive_free_cash_flow_flag"] = (
        pd.to_numeric(dataset["free_cash_flow_cr"], errors="coerce") > 0
    ).astype(int)
    dataset = add_revenue_cagr_3yr(dataset, profitandloss)
    dataset = add_fcf_cagr_5yr(dataset, cashflow)
    dataset = add_debt_to_equity_declining(dataset)
    dataset = calculate_composite_scores(dataset)
    return dataset


def latest_company_rows(df):
    if "year" not in df.columns:
        return df.copy()

    sortable = df.dropna(subset=["company_id", "year"]).copy()
    sortable["year"] = pd.to_numeric(sortable["year"], errors="coerce")
    return (
        sortable.sort_values(["company_id", "year"])
        .groupby("company_id", as_index=False)
        .tail(1)
        .reset_index(drop=True)
    )


def run_quality_compounder(data=None, db_path="db/nifty100.db"):
    return _run_preset("quality_compounder", data, db_path)


def run_value_pick(data=None, db_path="db/nifty100.db"):
    return _run_preset("value_pick", data, db_path)


def run_growth_accelerator(data=None, db_path="db/nifty100.db"):
    return _run_preset("growth_accelerator", data, db_path)


def run_dividend_champion(data=None, db_path="db/nifty100.db"):
    return _run_preset("dividend_champion", data, db_path)


def run_debt_free_blue_chip(data=None, db_path="db/nifty100.db"):
    return _run_preset("debt_free_blue_chip", data, db_path)


def run_turnaround_watch(data=None, db_path="db/nifty100.db"):
    dataset = _load_or_copy(data, db_path)
    dataset = add_debt_to_equity_declining(dataset)
    return ScreenerEngine().run_preset("turnaround_watch", latest_company_rows(dataset))


def add_revenue_cagr_3yr(df, profitandloss):
    result = df.copy()
    if "revenue_cagr_3yr" not in result.columns:
        result["revenue_cagr_3yr"] = pd.NA

    cagr_lookup = {}
    for company_id, group in profitandloss.groupby("company_id"):
        records = group.sort_values("year").to_dict("records")
        for _, row in group.iterrows():
            year = row["year"]
            historical_records = [
                record
                for record in records
                if record.get("year") is not None and record["year"] <= year
            ]
            cagr = calculate_metric_cagrs(
                historical_records,
                "sales",
                "revenue",
                periods=(3,),
            )
            cagr_lookup[(company_id, year)] = cagr["revenue_cagr_3yr"]

    result["revenue_cagr_3yr"] = result.apply(
        lambda row: cagr_lookup.get((row.get("company_id"), row.get("year")), row.get("revenue_cagr_3yr")),
        axis=1,
    )
    return result


def add_debt_to_equity_declining(df):
    result = df.copy()
    result["debt_to_equity_declining"] = False
    if not {"company_id", "year", "debt_to_equity"}.issubset(result.columns):
        return result

    result["_year_sort"] = pd.to_numeric(result["year"], errors="coerce")
    result["_de_sort"] = pd.to_numeric(result["debt_to_equity"], errors="coerce")
    result = result.sort_values(["company_id", "_year_sort"])
    previous = result.groupby("company_id")["_de_sort"].shift(1)
    result["debt_to_equity_declining"] = result["_de_sort"] < previous
    return result.drop(columns=["_year_sort", "_de_sort"]).sort_index()


def ensure_composite_quality_score(df):
    return calculate_composite_scores(df)


def calculate_composite_scores(df):
    result = df.copy()
    result["positive_free_cash_flow_flag"] = (
        pd.to_numeric(result.get("free_cash_flow_cr", 0), errors="coerce") > 0
    ).astype(int)

    global_scores = _weighted_scores(result)
    result["composite_quality_score"] = global_scores.clip(0, 100)

    if "broad_sector" not in result.columns:
        result["sector_relative_score"] = result["composite_quality_score"]
        return result

    sector_scores = pd.Series(index=result.index, dtype="float64")
    for _, sector_df in result.groupby(result["broad_sector"].fillna("UNKNOWN")):
        sector_scores.loc[sector_df.index] = _weighted_scores(sector_df).to_numpy(dtype="float64")

    result["sector_relative_score"] = sector_scores.fillna(0).clip(0, 100)
    return result


def winsorize_series(values):
    values = pd.to_numeric(values, errors="coerce").replace([float("inf"), -float("inf")], pd.NA)
    if values.dropna().empty:
        return values

    p10 = values.quantile(0.10)
    p90 = values.quantile(0.90)
    return values.clip(lower=p10, upper=p90)


def scale_metric(values, higher_is_better=True):
    winsorized = winsorize_series(values)
    if winsorized.dropna().empty:
        return pd.Series(0, index=values.index, dtype="float64")

    min_value = winsorized.min()
    max_value = winsorized.max()
    if pd.isna(min_value) or pd.isna(max_value) or min_value == max_value:
        return winsorized.notna().astype(float) * 50

    scaled = ((winsorized - min_value) / (max_value - min_value)) * 100
    if not higher_is_better:
        scaled = 100 - scaled

    return scaled.fillna(0).clip(0, 100).astype("float64")


def add_fcf_cagr_5yr(df, cashflow):
    result = df.copy()
    if "fcf_cagr_5yr" not in result.columns:
        result["fcf_cagr_5yr"] = pd.NA

    cashflow = cashflow.copy()
    cashflow["free_cash_flow_cr"] = (
        pd.to_numeric(cashflow["operating_activity"], errors="coerce")
        + pd.to_numeric(cashflow["investing_activity"], errors="coerce")
    )

    cagr_lookup = {}
    for company_id, group in cashflow.groupby("company_id"):
        records = group.sort_values("year").to_dict("records")
        for _, row in group.iterrows():
            year = row["year"]
            historical_records = [
                record
                for record in records
                if record.get("year") is not None and record["year"] <= year
            ]
            cagr = calculate_metric_cagrs(
                historical_records,
                "free_cash_flow_cr",
                "fcf",
                periods=(5,),
            )
            cagr_lookup[(company_id, year)] = cagr["fcf_cagr_5yr"]

    result["fcf_cagr_5yr"] = result.apply(
        lambda row: cagr_lookup.get((row.get("company_id"), row.get("year")), row.get("fcf_cagr_5yr")),
        axis=1,
    )
    return result


def export_preset_screeners_to_excel(
    db_path="db/nifty100.db",
    output_path="output/screener_output.xlsx",
):
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    data = load_screener_dataset(db_path)
    preset_functions = {
        "Quality Compounder": run_quality_compounder,
        "Value Pick": run_value_pick,
        "Growth Accelerator": run_growth_accelerator,
        "Dividend Champion": run_dividend_champion,
        "Debt-Free Blue Chip": run_debt_free_blue_chip,
        "Turnaround Watch": run_turnaround_watch,
    }

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, preset_function in preset_functions.items():
            result = preset_function(data).sort_values(
                "composite_quality_score",
                ascending=False,
                na_position="last",
            )
            export_df = _export_columns(result)
            export_df.to_excel(writer, sheet_name=sheet_name, index=False)

    _apply_conditional_formatting(output_path)
    return output_path


def _weighted_scores(df):
    total = pd.Series(0.0, index=df.index)
    for column, weight, higher_is_better in SCORING_METRICS:
        if column == "interest_coverage" and column in df.columns:
            values = _effective_interest_coverage(df)
        elif column in df.columns:
            values = pd.to_numeric(df[column], errors="coerce")
        else:
            values = pd.Series(pd.NA, index=df.index)

        total += scale_metric(values, higher_is_better).astype("float64") * weight

    return total.astype("float64").clip(0, 100)


def _cfo_pat_ratio(df):
    cfo = pd.to_numeric(df.get("operating_activity"), errors="coerce")
    pat = pd.to_numeric(df.get("net_profit"), errors="coerce")
    return cfo / pat.mask(pat == 0)


def _export_columns(df):
    output = pd.DataFrame()
    for source_column, display_column in EXPORT_COLUMNS:
        output[display_column] = df[source_column] if source_column in df.columns else pd.NA
    return output


def _apply_conditional_formatting(output_path):
    workbook = load_workbook(output_path)
    green = PatternFill(start_color=GREEN_FILL, end_color=GREEN_FILL, fill_type="solid")
    red = PatternFill(start_color=RED_FILL, end_color=RED_FILL, fill_type="solid")
    rules_by_sheet = {
        "Quality Compounder": {
            "ROE": (">", 15),
            "Debt-to-Equity": ("<", 1),
            "Free Cash Flow": (">", 0),
            "Revenue CAGR": (">", 10),
        },
        "Value Pick": {
            "P/E": ("<", 20),
            "P/B": ("<", 3),
            "Debt-to-Equity": ("<", 2),
            "Dividend Yield": (">", 1),
        },
        "Growth Accelerator": {
            "PAT CAGR": (">", 20),
            "Revenue CAGR": (">", 15),
            "Debt-to-Equity": ("<", 2),
        },
        "Dividend Champion": {
            "Dividend Yield": (">", 2),
            "Dividend Payout": ("<", 80),
            "Free Cash Flow": (">", 0),
        },
        "Debt-Free Blue Chip": {
            "Debt-to-Equity": ("=", 0),
            "ROE": (">", 12),
            "Sales": (">", 5000),
        },
        "Turnaround Watch": {
            "Revenue CAGR": (">", 10),
            "Free Cash Flow": (">", 0),
        },
    }

    for sheet_name, rules in rules_by_sheet.items():
        worksheet = workbook[sheet_name]
        headers = {cell.value: cell.column for cell in worksheet[1]}
        for header, rule in rules.items():
            if header not in headers:
                continue
            column_index = headers[header]
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=column_index)
                cell.fill = green if _cell_satisfies(cell.value, rule) else red

    workbook.save(output_path)


def _cell_satisfies(value, rule):
    operator, threshold = rule
    try:
        value = float(value)
    except (TypeError, ValueError):
        return False

    if operator == ">":
        return value > threshold
    if operator == "<":
        return value < threshold
    if operator == "=":
        return value == threshold
    return False


def _legacy_composite_quality_score(df):
    """Kept for reference during migration to weighted Day 17 scoring."""

    metric_columns = [
        "return_on_equity_pct",
        "free_cash_flow_cr",
        "revenue_cagr_5yr",
        "pat_cagr_5yr",
        "operating_profit_margin_pct",
        "interest_coverage",
        "asset_turnover",
    ]
    available = [column for column in metric_columns if column in df.columns]
    if not available:
        df["composite_quality_score"] = 0
        return df

    normalized = pd.DataFrame(index=df.index)
    for column in available:
        values = pd.to_numeric(df[column], errors="coerce")
        max_value = values.max()
        normalized[column] = 0 if pd.isna(max_value) or max_value == 0 else values / max_value

    df["composite_quality_score"] = normalized.mean(axis=1).fillna(0) * 100
    return df


def _run_preset(preset_name, data=None, db_path="db/nifty100.db"):
    dataset = latest_company_rows(_load_or_copy(data, db_path))
    return ScreenerEngine().run_preset(preset_name, dataset)


def _load_or_copy(data, db_path):
    if data is not None:
        return data.copy()
    return load_screener_dataset(db_path)


def _apply_debt_to_equity_max(df, column, threshold):
    values = pd.to_numeric(df[column], errors="coerce")
    financials = _is_financials(df)
    return df[financials | (values < threshold)]


def _apply_debt_to_equity_equal(df, threshold):
    if "debt_to_equity" not in df.columns:
        return df.iloc[0:0]

    values = pd.to_numeric(df["debt_to_equity"], errors="coerce")
    return df[values == threshold]


def _effective_interest_coverage(df):
    values = pd.to_numeric(df["interest_coverage"], errors="coerce")
    if "icr_label" not in df.columns:
        return values

    return values.mask(df["icr_label"].eq("Debt Free"), float("inf"))


def _is_financials(df):
    if "broad_sector" not in df.columns:
        return pd.Series(False, index=df.index)

    return df["broad_sector"].astype(str).str.strip().str.upper().eq(FINANCIAL_SECTOR)
