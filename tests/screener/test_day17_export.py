from pathlib import Path

from openpyxl import load_workbook
import pandas as pd

from src.screener.engine import (
    GREEN_FILL,
    calculate_composite_scores,
    export_preset_screeners_to_excel,
    scale_metric,
    winsorize_series,
)


def test_winsorisation_caps_percentiles():
    values = pd.Series([1, 2, 3, 100])

    result = winsorize_series(values)

    assert result.min() == values.quantile(0.10)
    assert result.max() == values.quantile(0.90)


def test_scale_metric_returns_zero_to_hundred_scores():
    result = scale_metric(pd.Series([10, 20, 30]))

    assert result.min() == 0
    assert result.max() == 100


def test_composite_score_calculation_is_bounded():
    scored = calculate_composite_scores(_score_data())

    assert scored["composite_quality_score"].between(0, 100).all()
    assert scored["sector_relative_score"].between(0, 100).all()


def test_sector_relative_score_is_generated_within_each_sector():
    scored = calculate_composite_scores(_score_data())

    assert "sector_relative_score" in scored.columns
    assert scored["sector_relative_score"].notna().all()
    assert scored[scored["broad_sector"].eq("Tech")]["sector_relative_score"].max() == 100


def test_excel_export_generation_and_worksheets(tmp_path):
    if not Path("db/nifty100.db").exists():
        return

    output_path = tmp_path / "screener_output.xlsx"

    export_preset_screeners_to_excel("db/nifty100.db", output_path)
    workbook = load_workbook(output_path)

    assert set(workbook.sheetnames) == {
        "Quality Compounder",
        "Value Pick",
        "Growth Accelerator",
        "Dividend Champion",
        "Debt-Free Blue Chip",
        "Turnaround Watch",
    }


def test_excel_export_is_sorted_by_composite_score(tmp_path):
    if not Path("db/nifty100.db").exists():
        return

    output_path = tmp_path / "screener_output.xlsx"

    export_preset_screeners_to_excel("db/nifty100.db", output_path)
    workbook = load_workbook(output_path, data_only=True)
    worksheet = workbook["Quality Compounder"]
    headers = {cell.value: cell.column for cell in worksheet[1]}
    score_column = headers["Composite Quality Score"]
    scores = [
        worksheet.cell(row=row, column=score_column).value
        for row in range(2, worksheet.max_row + 1)
    ]

    assert scores == sorted(scores, reverse=True)


def test_conditional_formatting_applies_green_to_passing_rule(tmp_path):
    if not Path("db/nifty100.db").exists():
        return

    output_path = tmp_path / "screener_output.xlsx"

    export_preset_screeners_to_excel("db/nifty100.db", output_path)
    workbook = load_workbook(output_path)
    worksheet = workbook["Quality Compounder"]
    headers = {cell.value: cell.column for cell in worksheet[1]}
    roe_cell = worksheet.cell(row=2, column=headers["ROE"])

    assert roe_cell.fill.fgColor.rgb.endswith(GREEN_FILL)


def _score_data():
    return pd.DataFrame(
        [
            {
                "company_id": "A",
                "broad_sector": "Tech",
                "return_on_equity_pct": 20,
                "return_on_capital_employed_pct": 18,
                "net_profit_margin_pct": 12,
                "fcf_cagr_5yr": 10,
                "cfo_pat_ratio": 1.2,
                "free_cash_flow_cr": 100,
                "revenue_cagr_5yr": 14,
                "pat_cagr_5yr": 16,
                "debt_to_equity": 0.4,
                "interest_coverage": 8,
            },
            {
                "company_id": "B",
                "broad_sector": "Tech",
                "return_on_equity_pct": 10,
                "return_on_capital_employed_pct": 8,
                "net_profit_margin_pct": 6,
                "fcf_cagr_5yr": 2,
                "cfo_pat_ratio": 0.6,
                "free_cash_flow_cr": -10,
                "revenue_cagr_5yr": 4,
                "pat_cagr_5yr": 3,
                "debt_to_equity": 2,
                "interest_coverage": 2,
            },
            {
                "company_id": "C",
                "broad_sector": "Financials",
                "return_on_equity_pct": 15,
                "return_on_capital_employed_pct": 11,
                "net_profit_margin_pct": 9,
                "fcf_cagr_5yr": 8,
                "cfo_pat_ratio": 1,
                "free_cash_flow_cr": 20,
                "revenue_cagr_5yr": 9,
                "pat_cagr_5yr": 11,
                "debt_to_equity": 5,
                "interest_coverage": 4,
            },
        ]
    )
