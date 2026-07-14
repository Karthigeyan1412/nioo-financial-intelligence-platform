import sqlite3

from openpyxl import load_workbook
import pandas as pd

from src.reports.peer_comparison import (
    BENCHMARK_FILL,
    METRIC_COLUMNS,
    generate_peer_comparison_report,
)


def test_peer_comparison_workbook_creation(tmp_path):
    db_path = _create_test_db(tmp_path)
    output_path = tmp_path / "peer_comparison.xlsx"

    result = generate_peer_comparison_report(db_path, output_path)

    assert result == output_path
    assert output_path.exists()
    assert output_path.stat().st_size > 0


def test_peer_comparison_has_exactly_eleven_worksheets(tmp_path):
    db_path = _create_test_db(tmp_path)
    output_path = tmp_path / "peer_comparison.xlsx"

    generate_peer_comparison_report(db_path, output_path)
    workbook = load_workbook(output_path)

    assert len(workbook.sheetnames) == 11


def test_required_columns_exist(tmp_path):
    db_path = _create_test_db(tmp_path)
    output_path = tmp_path / "peer_comparison.xlsx"

    generate_peer_comparison_report(db_path, output_path)
    worksheet = load_workbook(output_path)["Group 01"]
    headers = [cell.value for cell in worksheet[1]]

    assert "company_id" in headers
    assert "company_name" in headers
    for metric in METRIC_COLUMNS:
        assert metric in headers
        assert f"{metric}_percentile_rank" in headers


def test_median_summary_row_exists(tmp_path):
    db_path = _create_test_db(tmp_path)
    output_path = tmp_path / "peer_comparison.xlsx"

    generate_peer_comparison_report(db_path, output_path)
    worksheet = load_workbook(output_path, data_only=True)["Group 01"]

    assert worksheet.cell(row=worksheet.max_row, column=1).value == "MEDIAN"
    assert worksheet.cell(row=worksheet.max_row, column=2).value == "Peer Group Median"


def test_percentile_conditional_formatting_is_applied(tmp_path):
    db_path = _create_test_db(tmp_path)
    output_path = tmp_path / "peer_comparison.xlsx"

    generate_peer_comparison_report(db_path, output_path)
    worksheet = load_workbook(output_path)["Group 01"]

    assert len(worksheet.conditional_formatting) > 0


def test_benchmark_row_highlighting_exists(tmp_path):
    db_path = _create_test_db(tmp_path)
    output_path = tmp_path / "peer_comparison.xlsx"

    generate_peer_comparison_report(db_path, output_path)
    worksheet = load_workbook(output_path)["Group 01"]
    headers = {cell.value: cell.column for cell in worksheet[1]}

    benchmark_row = None
    for row in range(2, worksheet.max_row):
        if worksheet.cell(row=row, column=headers["company_id"]).value == "G01A":
            benchmark_row = row
            break

    assert benchmark_row is not None
    assert worksheet.cell(row=benchmark_row, column=1).fill.fgColor.rgb.endswith(BENCHMARK_FILL)


def _create_test_db(tmp_path):
    db_path = tmp_path / "test.db"
    peer_groups = []
    companies = []
    ratios = []
    percentiles = []

    for group_index in range(1, 12):
        peer_group_name = f"Group {group_index:02d}"
        for company_suffix, benchmark, multiplier in [("A", 1, 1), ("B", 0, 2)]:
            company_id = f"G{group_index:02d}{company_suffix}"
            peer_groups.append(
                {
                    "peer_group_name": peer_group_name,
                    "company_id": company_id,
                    "is_benchmark": benchmark,
                }
            )
            companies.append(
                {
                    "company_id": company_id,
                    "company_name": f"Company {company_id}",
                    "roce_percentage": 10 * multiplier,
                }
            )

            ratio_row = {
                "company_id": company_id,
                "year": 2025,
            }
            for metric in METRIC_COLUMNS:
                if metric == "return_on_capital_employed_pct":
                    continue
                ratio_row[metric] = float(group_index * multiplier)
            ratios.append(ratio_row)

            for metric in METRIC_COLUMNS[:3]:
                percentiles.append(
                    {
                        "company_id": company_id,
                        "peer_group_name": peer_group_name,
                        "metric": metric,
                        "value": float(group_index * multiplier),
                        "percentile_rank": 0.8 if benchmark else 0.2,
                        "year": 2025,
                    }
                )

    with sqlite3.connect(db_path) as connection:
        pd.DataFrame(peer_groups).to_sql("peer_groups", connection, index=False)
        pd.DataFrame(companies).to_sql("companies", connection, index=False)
        pd.DataFrame(ratios).to_sql("financial_ratios", connection, index=False)
        pd.DataFrame(percentiles).to_sql("peer_percentiles", connection, index=False)

    return db_path
